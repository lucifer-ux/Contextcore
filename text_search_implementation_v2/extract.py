# text_search_implementation_v2/extract.py

from pathlib import Path
import csv
import json
import subprocess
import zipfile
import xml.etree.ElementTree as ET

PLAIN_TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".json",
    ".xml",
    ".yaml",
    ".yml",
    ".toml",
    ".ini",
    ".cfg",
    ".conf",
    ".log",
    ".html",
    ".htm",
    ".rst",
    ".tsv",
}

ALLOWED_EXTENSIONS = {
    *PLAIN_TEXT_EXTENSIONS,
    ".pdf",
    ".doc",
    ".docx",
    ".ppt",
    ".pptx",
    ".xls",
    ".xlsx",
    ".ods",
    ".csv",
    ".rtf",
}


def _read_plain_text(path: Path) -> str:
    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return path.read_text(encoding=encoding, errors="ignore")
        except Exception:
            continue
    return path.read_text(errors="ignore")


def _read_csv_like(path: Path, delimiter: str) -> str:
    rows: list[str] = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        for row in reader:
            values = [str(cell) for cell in row if str(cell).strip()]
            if values:
                rows.append("\t".join(values))
    return "\n".join(rows).strip()


def _read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            values = [str(v) for v in row if v is not None and str(v).strip()]
            if values:
                lines.append("\t".join(values))
    return "\n".join(lines).strip()


def _read_xls(path: Path) -> str:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    lines: list[str] = []
    for sheet in wb.sheets():
        lines.append(f"# Sheet: {sheet.name}")
        for idx in range(sheet.nrows):
            values = [str(v) for v in sheet.row_values(idx) if str(v).strip()]
            if values:
                lines.append("\t".join(values))
    return "\n".join(lines).strip()


def _read_ods(path: Path) -> str:
    from odf.opendocument import load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    doc = load(str(path))
    lines: list[str] = []
    for table in doc.spreadsheet.getElementsByType(Table):
        table_name = table.getAttribute("name") or "Sheet"
        lines.append(f"# Sheet: {table_name}")
        for row in table.getElementsByType(TableRow):
            cells: list[str] = []
            for cell in row.getElementsByType(TableCell):
                parts: list[str] = []
                for para in cell.getElementsByType(P):
                    for node in para.childNodes:
                        value = getattr(node, "data", None)
                        if value:
                            parts.append(value)
                text_value = " ".join(part.strip() for part in parts if part.strip())
                if text_value:
                    cells.append(text_value)
            if cells:
                lines.append("\t".join(cells))
    return "\n".join(lines).strip()


def extract_text(path: Path) -> str:
    ext = path.suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        return ""

    try:
        if ext in PLAIN_TEXT_EXTENSIONS:
            raw = _read_plain_text(path)
            if ext == ".json":
                try:
                    return json.dumps(json.loads(raw), indent=2, ensure_ascii=False)
                except Exception:
                    return raw
            return raw

        if ext == ".pdf":
            try:
                return subprocess.check_output(
                    ["pdftotext", str(path), "-"],
                    stderr=subprocess.DEVNULL,
                ).decode(errors="ignore")
            except Exception:
                from pypdf import PdfReader

                reader = PdfReader(str(path))
                pages = [(p.extract_text() or "") for p in reader.pages]
                return "\n".join(pages).strip()

        if ext == ".pptx":
            try:
                texts: list[str] = []
                with zipfile.ZipFile(path, "r") as zf:
                    slide_names = sorted(
                        n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")
                    )
                    for name in slide_names:
                        xml_data = zf.read(name)
                        root = ET.fromstring(xml_data)
                        for node in root.iter():
                            if node.tag.endswith("}t") and node.text:
                                texts.append(node.text)
                return "\n".join(texts).strip()
            except Exception:
                return subprocess.check_output(
                    ["pandoc", str(path), "-t", "plain"],
                    stderr=subprocess.DEVNULL,
                ).decode(errors="ignore")

        if ext in {".doc", ".docx", ".ppt", ".rtf"}:
            return subprocess.check_output(
                ["pandoc", str(path), "-t", "plain"],
                stderr=subprocess.DEVNULL,
            ).decode(errors="ignore")

        if ext == ".csv":
            return _read_csv_like(path, delimiter=",")

        if ext == ".tsv":
            return _read_csv_like(path, delimiter="\t")

        if ext == ".xlsx":
            return _read_xlsx(path)

        if ext == ".xls":
            return _read_xls(path)

        if ext == ".ods":
            return _read_ods(path)

    except Exception:
        return ""

    return ""
